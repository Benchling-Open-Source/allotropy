from functools import partial

import openpyxl

from allotropy.allotrope.models.adm.plate_reader.rec._2024._06.plate_reader import (
    Model,
)
from allotropy.allotrope.schema_mappers.adm.plate_reader.rec._2024._06.plate_reader import (
    Data,
    Mapper,
)
from allotropy.named_file_contents import NamedFileContents
from allotropy.parsers.release_state import ReleaseState
from allotropy.parsers.tecan_magellan.constants import DISPLAY_NAME
from allotropy.parsers.tecan_magellan.tecan_magellan_reader import (
    TecanMagellanReader,
)
from allotropy.parsers.tecan_magellan.tecan_magellan_structure import (
    create_measurement_groups,
    create_metadata,
    MagellanMetadata,
)
from allotropy.parsers.utils.pandas import map_rows
from allotropy.parsers.vendor_parser import VendorParser


class TecanMagellanParser(VendorParser[Data, Model]):
    DISPLAY_NAME = DISPLAY_NAME
    RELEASE_STATE = ReleaseState.RECOMMENDED
    SUPPORTED_EXTENSIONS = TecanMagellanReader.SUPPORTED_EXTENSIONS
    SCHEMA_MAPPER = Mapper

    @classmethod
    def sniff(cls, named_file_contents: NamedFileContents) -> bool:
        try:
            wb = openpyxl.load_workbook(
                named_file_contents.get_bytes_stream(), read_only=True
            )
            if len(wb.sheetnames) != 1:
                wb.close()
                return False
            ws = wb[wb.sheetnames[0]]
            found_well_positions = False
            found_date_of_measurement = False
            for row in ws.iter_rows(max_row=50, values_only=True):
                for cell in row:
                    if isinstance(cell, str):
                        if "Well positions" in cell:
                            found_well_positions = True
                        if "Date of measurement" in cell:
                            found_date_of_measurement = True
                if found_well_positions or found_date_of_measurement:
                    break
            wb.close()
            return found_well_positions or found_date_of_measurement
        except Exception:
            return False

    def create_data(self, named_file_contents: NamedFileContents) -> Data:
        reader = TecanMagellanReader(named_file_contents)
        metadata = MagellanMetadata.create(reader.metadata_lines)
        well_count = len(reader.data)

        return Data(
            create_metadata(metadata, named_file_contents.original_file_path),
            map_rows(
                reader.data,
                partial(
                    create_measurement_groups, metadata=metadata, well_count=well_count
                ),
            ),
        )
