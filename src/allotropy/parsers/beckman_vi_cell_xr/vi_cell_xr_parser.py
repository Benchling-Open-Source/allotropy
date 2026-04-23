from __future__ import annotations

from xml.etree import ElementTree
import zipfile

import openpyxl

from allotropy.allotrope.models.adm.cell_counting.rec._2024._09.cell_counting import (
    Model,
)
from allotropy.allotrope.schema_mappers.adm.cell_counting.rec._2024._09.cell_counting import (
    Data,
    Mapper,
)
from allotropy.exceptions import AllotropeConversionError
from allotropy.named_file_contents import NamedFileContents
from allotropy.parsers.beckman_vi_cell_xr.vi_cell_xr_reader import create_reader_data
from allotropy.parsers.beckman_vi_cell_xr.vi_cell_xr_structure import (
    create_measurement_group,
    create_metadata,
)
from allotropy.parsers.release_state import ReleaseState
from allotropy.parsers.vendor_parser import VendorParser

_VI_CELL_MARKERS = ("Vi-CELL", "Vi-Cell")


class ViCellXRParser(VendorParser[Data, Model]):
    DISPLAY_NAME = "Beckman Coulter Vi-Cell XR"
    RELEASE_STATE = ReleaseState.RECOMMENDED
    SUPPORTED_EXTENSIONS = "txt,xls,xlsx"
    SCHEMA_MAPPER = Mapper

    @classmethod
    def _check_xlsx_via_zip(cls, named_file_contents: NamedFileContents) -> bool:
        """Fallback for xlsx files that openpyxl cannot open (e.g. corrupt styles)."""
        stream = named_file_contents.get_bytes_stream()
        with zipfile.ZipFile(stream) as zf:
            if "xl/sharedStrings.xml" not in zf.namelist():
                return False
            with zf.open("xl/sharedStrings.xml") as ss:
                tree = ElementTree.parse(ss)  # noqa: S314
                for elem in tree.iter():
                    if elem.text and any(
                        marker in elem.text for marker in _VI_CELL_MARKERS
                    ):
                        return True
        return False

    @classmethod
    def sniff(cls, named_file_contents: NamedFileContents) -> bool:
        try:
            if named_file_contents.extension == "txt":
                named_file_contents.contents.seek(0)
                for raw_line in named_file_contents.contents:
                    text = (
                        raw_line.decode("utf-8")
                        if isinstance(raw_line, bytes)
                        else raw_line
                    )
                    text = text.strip()
                    if text:
                        named_file_contents.contents.seek(0)
                        return any(marker in text for marker in _VI_CELL_MARKERS)
                named_file_contents.contents.seek(0)
                return False
            if named_file_contents.extension == "xlsx":
                try:
                    wb = openpyxl.load_workbook(
                        named_file_contents.get_bytes_stream(), read_only=True
                    )
                    ws = wb[wb.sheetnames[0]]
                    first_row = next(ws.iter_rows(max_row=1, values_only=True), None)
                    wb.close()
                    if first_row is None:
                        return False
                    first_cell = str(first_row[0]) if first_row[0] is not None else ""
                    return any(marker in first_cell for marker in _VI_CELL_MARKERS)
                except Exception:
                    # Fallback: read xlsx as zip for files with corrupt styles
                    named_file_contents.contents.seek(0)
                    return cls._check_xlsx_via_zip(named_file_contents)
            # xls: cannot use openpyxl, check using xlrd-style heuristic
            # Read first few bytes to check it's an Excel file, then trust extension
            named_file_contents.contents.seek(0)
            header = named_file_contents.contents.read(8)
            named_file_contents.contents.seek(0)
            if isinstance(header, str):
                header = header.encode("utf-8")
            # OLE2 compound document magic number (xls files)
            return header[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
        except Exception:
            return False

    def create_data(self, named_file_contents: NamedFileContents) -> Data:
        reader_data = create_reader_data(named_file_contents)
        if not reader_data.data:
            msg = "Cannot parse ASM from empty file."
            raise AllotropeConversionError(msg)

        return Data(
            create_metadata(reader_data, named_file_contents.original_file_path),
            [create_measurement_group(row) for row in reader_data.data],
        )
