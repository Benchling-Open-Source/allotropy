from __future__ import annotations

from dataclasses import dataclass
from io import StringIO
import re
from typing import Any, ClassVar

import openpyxl
import pandas as pd

from allotropy.named_file_contents import NamedFileContents
from allotropy.parsers.beckman_vi_cell_xr.constants import (
    DATE_HEADER,
    DEFAULT_VERSION,
    HEADINGS_TO_PARSER_HEADINGS,
    MODEL_RE,
    XrVersion,
)
from allotropy.parsers.lines_reader import read_to_lines
from allotropy.parsers.utils.pandas import (
    assert_value_from_df,
    df_to_series,
    read_csv,
    read_excel,
    SeriesData,
)
from allotropy.parsers.utils.values import assert_not_none


@dataclass
class ViCellData:
    data: list[SeriesData]
    serial_number: str | None
    version: XrVersion


def _is_report_format(named_file_contents: NamedFileContents) -> bool:
    """Detect the single-sample report format by checking for key-value layout."""
    if named_file_contents.extension not in ("xls", "xlsx"):
        return False
    try:
        wb = openpyxl.load_workbook(
            named_file_contents.get_bytes_stream(), read_only=True, data_only=True
        )
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))
        wb.close()
        if rows and rows[0][0] == "Sample ID":
            return True
    except Exception:
        return False
    return False


def create_reader_data(named_file_contents: NamedFileContents) -> ViCellData:
    if named_file_contents.extension == "txt":
        reader: ViCellXRReader | ViCellXRTXTReader | ViCellXRReportReader = (
            ViCellXRTXTReader(named_file_contents)
        )
    elif _is_report_format(named_file_contents):
        named_file_contents.contents.seek(0)
        reader = ViCellXRReportReader(named_file_contents)
    else:
        reader = ViCellXRReader(named_file_contents)
    return ViCellData(reader.data, reader.serial_number, reader.version)


def _get_file_version(version_str: str | None) -> XrVersion:
    match = re.match(MODEL_RE, version_str or "", flags=re.IGNORECASE)
    if not match:
        return DEFAULT_VERSION
    try:
        version = assert_not_none(match).groupdict()["version"]
        # TODO: raise exception for unsupported versions
        version = ".".join(version.split(".")[0:2])
        return XrVersion(version)
    except AttributeError:
        return DEFAULT_VERSION


class ViCellXRReader:
    data: list[SeriesData]
    serial_number: str | None
    version: XrVersion

    def __init__(self, named_file_contents: NamedFileContents) -> None:
        # calamine is faster for reading xlsx, but does not read xls. For xls, let pandas pick engine.
        self.engine = "calamine" if named_file_contents.extension == "xlsx" else None
        self.contents = named_file_contents.contents
        file_info = self._get_file_info()
        try:
            self.serial_number = (
                file_info.get(str, "serial", "").split(":", maxsplit=1)[1].strip()
            )
        except (IndexError, ValueError):
            self.serial_number = None
        self.version = _get_file_version(file_info.get(str, "model"))
        self.data = self._read_data()

    def _read_excel(self, **kwargs: Any) -> pd.DataFrame:
        return read_excel(self.contents, engine=self.engine, **kwargs)

    def _read_data(self) -> list[SeriesData]:
        header_row = 4
        if self.version == XrVersion._2_04:
            header_row = 3

        header = self._get_file_header(header_row)
        skiprows = header_row + 1

        file_data = self._read_excel(skiprows=skiprows, names=header)

        # rename the columns to match the existing parser that was based on xls(x) files
        file_data = file_data.rename(columns=HEADINGS_TO_PARSER_HEADINGS)

        # Do the datetime conversion and remove all rows that fail to pass as datetime
        # This fixes an issue where some files have a hidden invalid first row
        file_data[DATE_HEADER] = pd.to_datetime(
            assert_value_from_df(file_data, DATE_HEADER),
            format="%d %b %Y  %I:%M:%S %p",
            errors="coerce",
        )
        file_data = file_data.dropna(subset=DATE_HEADER)
        return [SeriesData(x) for _, x in file_data.iterrows()]

    def _get_file_header(self, header_row: int) -> list[str]:
        """Combine the two rows that forms the header."""
        header = (
            self._read_excel(
                nrows=2,
                skiprows=header_row,
                header=None,
            )
            .fillna("")
            .astype(str)
        )

        header_list: list[str] = header.agg(
            lambda x: " ".join(x).replace(" /ml", "/ml").strip()
        ).to_list()
        return header_list

    def _get_file_info(self) -> SeriesData:
        info: pd.Series[Any] = self._read_excel(
            nrows=3,
            header=None,
            usecols=[0],
        ).squeeze()
        info.index = pd.Index(["model", "filepath", "serial"])
        return SeriesData(info)


class ViCellXRReportReader:
    """Reader for the single-sample report format exported by Vi-CELL XR 2.04."""

    data: list[SeriesData]
    serial_number: str | None
    version: XrVersion

    RESULTS_FIELDS: ClassVar[dict[str, int]] = {
        "Total cells": 9,
        "Viable cells": 10,
        "Viability (%)": 11,
        "Total cells/ml (x10^6)": 12,
        "Viable cells/ml (x10^6)": 13,
        "Avg. diam. (microns)": 14,
        "Avg. circ.": 15,
        "Images": 16,
        "Average cells / image": 17,
        "Avg. background intensity": 18,
    }

    SETTINGS_FIELDS: ClassVar[dict[str, int]] = {
        "Cell type": 9,
        "Minimum diameter (microns)": 10,
        "Maximum diameter (microns)": 11,
        "Minimum circularity": 12,
        "Dilution factor": 13,
        "Cell brightness (%)": 14,
        "Cell sharpness": 15,
        "Viable cell spot brightness (%)": 16,
        "Viable cell spot area (%)": 17,
        "Decluster degree": 18,
        "Aspirate cycles": 19,
        "Trypan blue mixing cycles": 20,
    }

    def __init__(self, named_file_contents: NamedFileContents) -> None:
        wb = openpyxl.load_workbook(
            named_file_contents.get_bytes_stream(), read_only=True, data_only=True
        )
        ws = wb[wb.sheetnames[0]]
        self.rows = list(ws.iter_rows(values_only=True))
        wb.close()

        self.version = _get_file_version(str(self.rows[0][0]))
        self.serial_number = None
        self.data = self._read_data()

    def _read_data(self) -> list[SeriesData]:
        data: dict[str, Any] = {}

        data["Sample ID"] = self.rows[3][2]
        data["File name"] = self.rows[4][2]
        data[DATE_HEADER] = self.rows[5][2]
        comment = self.rows[6][2] if len(self.rows) > 6 else None
        if comment:
            data["Comment"] = comment

        for field, row_idx in self.RESULTS_FIELDS.items():
            if row_idx < len(self.rows):
                val = self.rows[row_idx][3]
                if val is not None:
                    data[field] = val

        for field, row_idx in self.SETTINGS_FIELDS.items():
            if row_idx < len(self.rows):
                val = self.rows[row_idx][8]
                if val is not None:
                    data[field] = val

        series = pd.Series(data)
        series[DATE_HEADER] = pd.to_datetime(
            series[DATE_HEADER],
            format="%d %b %Y  %I:%M:%S %p",
        )
        return [SeriesData(series)]


class ViCellXRTXTReader:
    data: list[SeriesData]
    serial_number: str | None
    version: XrVersion

    def __init__(self, contents: NamedFileContents) -> None:
        self.lines = read_to_lines(contents)
        self.data = self._read_data()
        self.serial_number = self.data[0].get(str, "Unit S/N")
        self.version = _get_file_version(str(self.data[0].series.name))

    def _read_data(self) -> list[SeriesData]:
        data_frame = read_csv(
            StringIO("\n".join(self.lines)), sep=" :", index_col=0, engine="python"
        )
        file_data = df_to_series(data_frame.astype(str).T)
        file_data = file_data.str.strip().replace("", None)
        file_data = file_data.rename(index=HEADINGS_TO_PARSER_HEADINGS)
        # Do the datetime conversion and remove all rows that fail to pass as datetime
        # This fixes an issue where some files have a hidden invalid first row
        file_data[DATE_HEADER] = pd.to_datetime(
            file_data[DATE_HEADER],
            format="%d %b %Y  %I:%M:%S %p",
            errors="coerce",
        )
        return [SeriesData(file_data)]
