import openpyxl

from allotropy.allotrope.models.adm.cell_counting.rec._2024._09.cell_counting import (
    Model,
)
from allotropy.allotrope.schema_mappers.adm.cell_counting.rec._2024._09.cell_counting import (
    Data,
    Mapper,
)
from allotropy.named_file_contents import NamedFileContents
from allotropy.parsers.release_state import ReleaseState
from allotropy.parsers.roche_cedex_hires.constants import DISPLAY_NAME
from allotropy.parsers.roche_cedex_hires.roche_cedex_hires_reader import (
    RocheCedexHiResReader,
)
from allotropy.parsers.roche_cedex_hires.roche_cedex_hires_structure import (
    create_measurement_groups,
    create_metadata,
)
from allotropy.parsers.utils.pandas import map_rows
from allotropy.parsers.vendor_parser import VendorParser


class RocheCedexHiResParser(VendorParser[Data, Model]):
    DISPLAY_NAME = DISPLAY_NAME
    RELEASE_STATE = ReleaseState.RECOMMENDED
    SUPPORTED_EXTENSIONS = RocheCedexHiResReader.SUPPORTED_EXTENSIONS
    SCHEMA_MAPPER = Mapper

    @classmethod
    def sniff(cls, named_file_contents: NamedFileContents) -> bool:
        try:
            if named_file_contents.extension == "xlsx":
                wb = openpyxl.load_workbook(
                    named_file_contents.get_bytes_stream(), read_only=True
                )
                ws = wb[wb.sheetnames[0]]
                first_row = next(ws.iter_rows(max_row=1, values_only=True), None)
                wb.close()
                if first_row is None:
                    return False
                cells = [str(c) for c in first_row if c is not None]
                return any("Cedex ID" in c or "identifer" in c for c in cells)
            named_file_contents.contents.seek(0)
            raw = named_file_contents.contents.read(8192)
            text = (
                raw.decode("utf-8", errors="replace") if isinstance(raw, bytes) else raw
            )
            lines = text.splitlines()
            if not lines:
                return False
            header = lines[0]
            return "Cedex ID" in header or "identifer" in header
        except Exception:
            return False

    def create_data(self, named_file_contents: NamedFileContents) -> Data:
        reader = RocheCedexHiResReader(named_file_contents)
        return Data(
            create_metadata(reader.header, named_file_contents.original_file_path),
            map_rows(reader.data, create_measurement_groups),
        )
