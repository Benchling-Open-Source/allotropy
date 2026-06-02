import openpyxl

from allotropy.allotrope.models.adm.plate_reader.rec._2025._03.plate_reader import (
    Model,
)
from allotropy.allotrope.schema_mappers.adm.plate_reader.rec._2025._03.plate_reader import (
    Data,
    Mapper,
)
from allotropy.named_file_contents import NamedFileContents
from allotropy.parsers.release_state import ReleaseState
from allotropy.parsers.thermo_skanit.thermo_skanit_structure import DataThermoSkanIt
from allotropy.parsers.utils.pandas import read_multisheet_excel
from allotropy.parsers.vendor_parser import VendorParser


class ThermoSkanItParser(VendorParser[Data, Model]):
    DISPLAY_NAME = "Thermo Fisher Scientific SkanIt"
    RELEASE_STATE = ReleaseState.RECOMMENDED
    SUPPORTED_EXTENSIONS = "xlsx"
    SUPPORTED_DETECTION_MODES = "Absorbance, Fluorescence, Luminescence"
    SCHEMA_MAPPER = Mapper

    @classmethod
    def sniff(cls, named_file_contents: NamedFileContents) -> bool:
        try:
            wb = openpyxl.load_workbook(
                named_file_contents.get_bytes_stream(), read_only=True
            )
            sheet_names = set(wb.sheetnames)
            # Full export with session/instrument sheets
            if (
                "Session information" in sheet_names
                and "Instrument information" in sheet_names
            ):
                wb.close()
                return True
            # Simplified export: first cell is "Measurement results"
            ws = wb[wb.sheetnames[0]]
            first_row = next(ws.iter_rows(max_row=1, values_only=True), None)
            wb.close()
            if first_row and first_row[0] == "Measurement results":
                return True
            return False
        except Exception:
            return False

    def create_data(self, named_file_contents: NamedFileContents) -> Data:
        contents = read_multisheet_excel(
            named_file_contents.contents, engine="calamine"
        )
        return DataThermoSkanIt.create(
            sheet_data=contents, file_path=named_file_contents.original_file_path
        )
