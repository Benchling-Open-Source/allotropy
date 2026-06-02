import openpyxl

from allotropy.allotrope.models.adm.spectrophotometry.benchling._2023._12.spectrophotometry import (
    Model,
)
from allotropy.allotrope.schema_mappers.adm.spectrophotometry.benchling._2023._12.spectrophotometry import (
    Data,
    Mapper,
)
from allotropy.named_file_contents import NamedFileContents
from allotropy.parsers.release_state import ReleaseState
from allotropy.parsers.thermo_fisher_qubit4.thermo_fisher_qubit4_reader import (
    ThermoFisherQubit4Reader,
)
from allotropy.parsers.thermo_fisher_qubit_flex import constants
from allotropy.parsers.thermo_fisher_qubit_flex.thermo_fisher_qubit_flex_reader import (
    ThermoFisherQubitFlexReader,
)
from allotropy.parsers.thermo_fisher_qubit_flex.thermo_fisher_qubit_flex_structure import (
    create_data,
)
from allotropy.parsers.vendor_parser import VendorParser


class ThermoFisherQubitFlexParser(VendorParser[Data, Model]):
    DISPLAY_NAME = constants.DISPLAY_NAME
    RELEASE_STATE = ReleaseState.RECOMMENDED
    SUPPORTED_EXTENSIONS = ThermoFisherQubit4Reader.SUPPORTED_EXTENSIONS
    SUPPORTED_DETECTION_MODES = "Fluorescence"
    SCHEMA_MAPPER = Mapper

    @classmethod
    def sniff(cls, named_file_contents: NamedFileContents) -> bool:
        try:
            if named_file_contents.extension == "xlsx":
                wb = openpyxl.load_workbook(
                    named_file_contents.get_bytes_stream(), read_only=True
                )
                ws = wb[wb.sheetnames[0]]
                first_row = next(ws.iter_rows(max_row=1, values_only=True), None)
                wb.close()
                if first_row is None:
                    return False
                cells = [str(c) for c in first_row if c is not None]
                return "Test Date" in cells and "Sample RFU" in cells
            named_file_contents.contents.seek(0)
            raw = named_file_contents.contents.read(8192)
            text = (
                raw.decode("utf-8", errors="replace") if isinstance(raw, bytes) else raw
            )
            lines = text.splitlines()
            if not lines:
                return False
            # Skip sep= line if present
            header = lines[1] if lines[0].strip().startswith("sep=") else lines[0]
            return "Test Date" in header and (
                "Sample RFU" in header or "Run ID" in header
            )
        except Exception:
            return False

    def create_data(self, named_file_contents: NamedFileContents) -> Data:
        df = ThermoFisherQubitFlexReader.read(named_file_contents)
        return create_data(df, named_file_contents.original_file_path)
