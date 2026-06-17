from io import BytesIO

import openpyxl

from allotropy.allotrope.models.adm.solution_analyzer.rec._2024._09.solution_analyzer import (
    Model,
)
from allotropy.allotrope.schema_mappers.adm.solution_analyzer.rec._2024._09.solution_analyzer import (
    Data,
    Mapper,
)
from allotropy.named_file_contents import NamedFileContents
from allotropy.parsers.beckman_pharmspec.beckman_pharmspec_reader import (
    BeckmanPharmspecReader,
)
from allotropy.parsers.beckman_pharmspec.beckman_pharmspec_structure import (
    create_calculated_data,
    create_measurement_groups,
    create_metadata,
    Distribution,
    Header,
)
from allotropy.parsers.release_state import ReleaseState
from allotropy.parsers.utils.pandas import read_excel
from allotropy.parsers.vendor_parser import VendorParser


class PharmSpecParser(VendorParser[Data, Model]):
    DISPLAY_NAME = "Beckman Coulter PharmSpec"
    RELEASE_STATE = ReleaseState.RECOMMENDED
    SUPPORTED_EXTENSIONS = BeckmanPharmspecReader.SUPPORTED_EXTENSIONS
    SUPPORTED_DETECTION_MODES = "Light Obscuration"
    SCHEMA_MAPPER = Mapper

    @classmethod
    def sniff(cls, named_file_contents: NamedFileContents) -> bool:
        if named_file_contents.extension == "xls":
            return cls._sniff_xls(named_file_contents)
        try:
            wb = openpyxl.load_workbook(
                named_file_contents.get_bytes_stream(), read_only=True
            )
            ws = wb[wb.sheetnames[0]]
            for row in ws.iter_rows(max_row=20, values_only=True):
                for cell in row:
                    if isinstance(cell, str) and "Particle" in cell:
                        wb.close()
                        return True
            wb.close()
            return False
        except Exception:
            return False

    @classmethod
    def _sniff_xls(cls, named_file_contents: NamedFileContents) -> bool:
        try:
            raw_content = named_file_contents.contents.read()
            named_file_contents.contents.seek(0)
            content_bytes = (
                raw_content.encode("utf-8")
                if isinstance(raw_content, str)
                else raw_content
            )
            # HTML-based XLS files can be sniffed as text
            if content_bytes.lstrip()[:6].lower().startswith((b"<html>", b"<html ")):
                text = content_bytes[:4000].decode("utf-8", errors="ignore")
                return "Run Counter Test" in text or "Particle Size" in text
            # Binary XLS - read with pandas to check content
            df = read_excel(
                BytesIO(content_bytes),
                header=None,
                engine="calamine",
                nrows=20,
            )
            for _, row in df.iterrows():
                for cell in row.values:
                    if isinstance(cell, str) and (
                        "Run Counter Test" in cell or "Particle Size" in cell
                    ):
                        return True
            return False
        except Exception:
            return False

    def create_data(self, named_file_contents: NamedFileContents) -> Data:
        reader = BeckmanPharmspecReader(named_file_contents)
        distributions = Distribution.create_distributions(reader.data)
        header = Header.create(reader.header)

        return Data(
            create_metadata(header, named_file_contents.original_file_path),
            create_measurement_groups(header, distributions),
            create_calculated_data(distributions),
        )
