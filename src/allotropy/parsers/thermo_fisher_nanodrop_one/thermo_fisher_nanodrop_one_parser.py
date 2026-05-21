from pathlib import Path

import openpyxl
import pandas as pd

from allotropy.allotrope.models.adm.spectrophotometry.benchling._2023._12.spectrophotometry import (
    Model,
)
from allotropy.allotrope.schema_mappers.adm.spectrophotometry.benchling._2023._12.spectrophotometry import (
    Data,
    Mapper,
)
from allotropy.constants import DEFAULT_ENCODING
from allotropy.named_file_contents import NamedFileContents
from allotropy.parsers.release_state import ReleaseState
from allotropy.parsers.thermo_fisher_nanodrop_one.constants import DISPLAY_NAME
from allotropy.parsers.thermo_fisher_nanodrop_one.thermo_fisher_nanodrop_one_structure import (
    DataNanodrop,
)
from allotropy.parsers.utils.pandas import read_csv, read_multisheet_excel
from allotropy.parsers.vendor_parser import VendorParser

_NANODROP_EXPERIMENT_PREFIXES = ("Nucleic Acid", "Protein A280", "Protein & Label")
_NANODROP_SHEET_PREFIXES = (
    "Nucleic Acid",
    "Protein A280",
    "Protein & Label",
    "dsDNA",
    "ssDNA",
    "RNA",
)


class ThermoFisherNanodropOneParser(VendorParser[Data, Model]):
    DISPLAY_NAME = DISPLAY_NAME
    RELEASE_STATE = ReleaseState.RECOMMENDED
    SUPPORTED_EXTENSIONS = "csv,xlsx"
    SCHEMA_MAPPER = Mapper

    @classmethod
    def sniff(cls, named_file_contents: NamedFileContents) -> bool:
        try:
            if named_file_contents.extension == "xlsx":
                wb = openpyxl.load_workbook(
                    named_file_contents.get_bytes_stream(), read_only=True
                )
                sheet_name = wb.sheetnames[0]
                # Check sheet name for known experiment type prefixes
                if any(
                    sheet_name.startswith(prefix) for prefix in _NANODROP_SHEET_PREFIXES
                ):
                    wb.close()
                    return True
                # Also check first row for NanodropOne-specific columns
                ws = wb[sheet_name]
                first_row = next(ws.iter_rows(max_row=1, values_only=True), None)
                wb.close()
                if first_row is None:
                    return False
                cells = [str(c) for c in first_row if c is not None]
                return "A260/A280" in cells and any("Nucleic Acid" in c for c in cells)
            named_file_contents.contents.seek(0)
            raw = named_file_contents.contents.read(8192)
            text = (
                raw.decode("utf-8", errors="replace") if isinstance(raw, bytes) else raw
            )
            lines = text.splitlines()
            if not lines:
                return False
            header = lines[0]
            return any(prefix in header for prefix in _NANODROP_EXPERIMENT_PREFIXES)
        except Exception:
            return False

    def read_data(
        self, named_file_contents: NamedFileContents
    ) -> tuple[pd.DataFrame, str]:
        if named_file_contents.extension == "xlsx":
            xlsx_data = read_multisheet_excel(
                named_file_contents.contents,
                engine="calamine",
            )
            sheet_name = next(iter(xlsx_data.keys()))
            experiment_type, *_ = sheet_name.split(maxsplit=1)
            return xlsx_data[sheet_name], experiment_type

        csv_data = read_csv(
            named_file_contents.contents,
            encoding=DEFAULT_ENCODING,
        )
        file_name = Path(named_file_contents.original_file_path).name
        experiment_type, *_ = file_name.split(maxsplit=1)
        return csv_data, experiment_type

    def create_data(self, named_file_contents: NamedFileContents) -> Data:
        data, experiment_type = self.read_data(named_file_contents)
        return DataNanodrop.create(
            data, experiment_type, named_file_contents.original_file_path
        )
