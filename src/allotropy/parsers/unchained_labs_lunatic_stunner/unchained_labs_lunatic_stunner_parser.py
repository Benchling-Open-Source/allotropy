from __future__ import annotations

import openpyxl

from allotropy.allotrope.models.adm.plate_reader.rec._2025._03.plate_reader import (
    Model,
)
from allotropy.allotrope.schema_mappers.adm.plate_reader.rec._2025._03.plate_reader import (
    Data,
    Mapper,
)
from allotropy.named_file_contents import NamedFileContents
from allotropy.parsers.release_state import ReleaseState
from allotropy.parsers.unchained_labs_lunatic_stunner.unchained_labs_lunatic_stunner_reader import (
    UnchainedLabsLunaticReader,
)
from allotropy.parsers.unchained_labs_lunatic_stunner.unchained_labs_lunatic_stunner_structure import (
    create_measurement_groups,
    create_metadata,
)
from allotropy.parsers.vendor_parser import VendorParser


class UnchainedLabsLunaticStunnerParser(VendorParser[Data, Model]):
    DISPLAY_NAME = "Unchained Labs Lunatic & Stunner"
    RELEASE_STATE = ReleaseState.RECOMMENDED
    SUPPORTED_EXTENSIONS = UnchainedLabsLunaticReader.SUPPORTED_EXTENSIONS
    SCHEMA_MAPPER = Mapper

    @classmethod
    def sniff(cls, named_file_contents: NamedFileContents) -> bool:
        try:
            if named_file_contents.extension == "xlsx":
                named_file_contents.contents.seek(0)
                wb = openpyxl.load_workbook(
                    named_file_contents.get_bytes_stream(), read_only=True
                )
                ws = wb[wb.sheetnames[0]]
                found = False
                for row in ws.iter_rows(max_row=50, values_only=True):
                    for cell in row:
                        if isinstance(cell, str) and "sample name" in cell.lower():
                            found = True
                            break
                    if found:
                        break
                wb.close()
                return found
            named_file_contents.contents.seek(0)
            raw = named_file_contents.contents.read(8192)
            text = (
                raw.decode("utf-8", errors="replace") if isinstance(raw, bytes) else raw
            )
            # Check the first portion of text as a block rather than per-line,
            # because CSV headers may contain quoted fields with embedded newlines
            block = text[:4096].lower()
            return "sample name" in block and (
                "plate id" in block or "plate type" in block
            )
        except Exception:
            return False

    def create_data(self, named_file_contents: NamedFileContents) -> Data:
        reader = UnchainedLabsLunaticReader(named_file_contents)
        measurement_groups, calculated_data = create_measurement_groups(
            reader.header, reader.data
        )
        return Data(
            create_metadata(reader.header, named_file_contents.original_file_path),
            measurement_groups=measurement_groups,
            calculated_data=calculated_data,
        )
