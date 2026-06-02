from functools import partial

import openpyxl

from allotropy.allotrope.models.adm.cell_counting.rec._2024._09.cell_counting import (
    Model,
)
from allotropy.allotrope.schema_mappers.adm.cell_counting.rec._2024._09.cell_counting import (
    Data,
    Mapper,
)
from allotropy.named_file_contents import NamedFileContents
from allotropy.parsers.release_state import ReleaseState
from allotropy.parsers.revvity_matrix.constants import DISPLAY_NAME
from allotropy.parsers.revvity_matrix.revvity_matrix_reader import (
    RevvityMatrixReader,
)
from allotropy.parsers.revvity_matrix.revvity_matrix_structure import (
    create_measurement_group,
    create_metadata,
)
from allotropy.parsers.utils.pandas import map_rows
from allotropy.parsers.vendor_parser import VendorParser

_REVVITY_CSV_COLUMNS = {"Row", "Column"}
_REVVITY_XLSX_COLUMNS = {"Well Name", "Live Count", "Viability"}


class RevvityMatrixParser(VendorParser[Data, Model]):
    DISPLAY_NAME = DISPLAY_NAME
    RELEASE_STATE = ReleaseState.RECOMMENDED
    SUPPORTED_EXTENSIONS = RevvityMatrixReader.SUPPORTED_EXTENSIONS
    SUPPORTED_DETECTION_MODES = "Brightfield"
    SCHEMA_MAPPER = Mapper

    @classmethod
    def sniff(cls, named_file_contents: NamedFileContents) -> bool:
        try:
            if named_file_contents.extension == "xlsx":
                wb = openpyxl.load_workbook(
                    named_file_contents.get_bytes_stream(), read_only=True
                )
                ws = wb[wb.sheetnames[0]]
                # Check for header row with Well Name, or metadata rows like "Plate Name"
                for row in ws.iter_rows(max_row=15, values_only=True):
                    cells = [str(c) for c in row if c is not None]
                    if any("Well Name" in c for c in cells) and any(
                        "Live Count" in c or "Live Concentration" in c for c in cells
                    ):
                        wb.close()
                        return True
                    # CSV-style: first row has "Row" and "Column"
                    if "Row" in cells and "Column" in cells:
                        wb.close()
                        return True
                wb.close()
                return False
            named_file_contents.contents.seek(0)
            raw = named_file_contents.contents.read(8192)
            text = (
                raw.decode("utf-8", errors="replace") if isinstance(raw, bytes) else raw
            )
            lines = text.splitlines()
            if not lines:
                return False
            first = lines[0]
            # Metadata-style CSV: starts with "Plate Name,..."
            if first.startswith("Plate Name,"):
                return True
            # Flat CSV: header row has "Well Name" and cell counting columns
            if "Well Name" in first and (
                "Live Count" in first or "Live Cells/mL" in first
            ):
                return True
            return False
        except Exception:
            return False

    def create_data(self, named_file_contents: NamedFileContents) -> Data:
        reader = RevvityMatrixReader(named_file_contents)
        return Data(
            create_metadata(named_file_contents.original_file_path, reader.headers),
            map_rows(
                reader.data, partial(create_measurement_group, headers=reader.headers)
            ),
        )
